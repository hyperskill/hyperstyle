import logging.config
from pathlib import Path
from typing import Union

import pandas as pd
from openpyxl import load_workbook, Workbook

logger = logging.getLogger(__name__)


def remove_sheet(workbook_path: Union[str, Path], sheet_name: str, to_raise_error: bool = False) -> None:
    try:
        workbook = load_workbook(workbook_path)
        workbook.remove(workbook[sheet_name])
        workbook.save(workbook_path)

    except KeyError as e:
        message = f'Sheet with specified name: {sheet_name} does not exist.'
        if to_raise_error:
            logger.exception(message)
            raise e
        else:
            logger.info(message)


def create_workbook(output_file_path: Path) -> Workbook:
    workbook = Workbook()
    workbook.save(output_file_path)
    return workbook


def write_dataframe_to_xlsx_sheet(xlsx_file_path: Union[str, Path], df: pd.DataFrame, sheet_name: str,
                                  mode: str = 'a', to_write_row_names: bool = False) -> None:
    """
    mode: str Available values are {'w', 'a'}. File mode to use (write or append).
    to_write_row_names: bool Write row names.
    """

    with pd.ExcelWriter(xlsx_file_path, mode=mode) as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=to_write_row_names)
